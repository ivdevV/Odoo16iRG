#!/usr/bin/env python3
"""
Analizador de acciones planificadas ir.cron para Odoo 16 IRG/ISEP.
Estrategia eficiente: construir índices globales primero, luego aplicar matching en memoria.
"""
import openpyxl, json, os, subprocess, re, sys
from datetime import datetime

BASE_DIR = '/Users/ivrogo/Workspace/Proyectos iRG/Odoo16iRG'
ADDONS_ROOT = os.path.join(BASE_DIR, 'addons-extra')
OUTPUT_JSON = os.path.join(BASE_DIR, 'doc/csv-analizer/_findings_crons.json')
OUTPUT_MD = os.path.join(BASE_DIR, 'doc/csv-analizer/_findings_crons.md')
XLSX_PATH = os.path.join(BASE_DIR, 'doc/csv-analizer/Acciones planificadas (ir.cron).xlsx')

# ============================================================
# CATEGORÍAS CONOCIDAS
# ============================================================
CUSTOM_CATEGORIES = {'extrairg', 'addons_uisep', 'addons_irg'}
OCA_CATEGORIES = {
    'account_financial_tools', 'account_related15', 'addons-extend',
    'community-16', 'addons-co', 'addons-mx', 'reporting-engine',
    'localizacion_espanola', 'edi-framework', 'vztech', 'pos',
}

# ============================================================
# CRONS NATIVOS ODOO — mapa de palabras clave a módulo
# ============================================================
NATIVE_KEYWORDS = [
    # (keyword, module) — orden de prioridad descendente (más específico primero)
    ('ludificación', 'gamification'),
    ('consolidación del seguimiento de karma', 'gamification'),
    ('gamification', 'gamification'),
    ('karma', 'gamification'),
    ('digest', 'digest'),
    ('resumen periódico', 'digest'),
    ('resumen de correo', 'digest'),
    ('correos electrónicos del resumen', 'digest'),
    ('enviar correos electrónicos', 'mail'),
    ('obtener correos electrónicos', 'mail'),
    ('limpiar correos', 'mail'),
    ('notificaciones push', 'mail'),
    ('cancelar notificaciones', 'mail'),
    ('gerente de la cola de correo electrónico', 'mail'),
    ('correo: gerente', 'mail'),
    ('notificación: eliminar notificaciones', 'mail'),
    ('notificaciones de mensajes', 'mail'),
    ('notificación: enviar notificaciones', 'mail'),
    ('mensajes programados', 'mail'),
    ('mailing', 'mass_mailing'),
    ('mail mass', 'mass_mailing'),
    ('correo masivo', 'mass_mailing'),
    ('campaña', 'mass_mailing'),
    ('sms', 'sms'),
    ('enviar sms', 'sms'),
    ('carta postal', 'snailmail'),
    ('snailmail', 'snailmail'),
    ('mapa del sitio', 'website'),
    ('sitemap', 'website'),
    ('deshabilitar snippets', 'website'),
    ('visitante del sitio web', 'website'),
    ('cesta abandonada', 'website_sale'),
    ('comercio electrónico', 'website_sale'),
    ('chat en vivo', 'im_livechat'),
    ('livechat', 'im_livechat'),
    ('actualizar divisa', 'base'),
    ('tipo de cambio', 'currency_rate_update'),
    ('actualizar tipo de cambio', 'currency_rate_update'),
    ('datos de moneda', 'base'),
    ('ejecutar reglas de automatización', 'base_automation'),
    ('norma de acción básica', 'base_automation'),
    ('acción automatizada', 'base_automation'),
    ('regla de automatización', 'base_automation'),
    ('módulos nuevos', 'base'),
    ('actualizaciones del módulo', 'base'),
    ('base: limpieza automática', 'base'),
    ('limpieza automática de datos', 'base'),
    ('depuración de datos', 'base_setup'),
    ('reciclado de datos', 'data_recycle'),
    ('fusión de datos', 'data_merge'),
    ('deduplicación', 'data_merge'),
    ('gestión de actividades', 'mail'),
    ('actividades vencidas', 'mail'),
    ('generar asientos contables', 'account'),
    ('asiento contable', 'account'),
    ('asientos contables pendientes', 'account'),
    ('facturas vencidas', 'account'),
    ('cobros diferidos', 'account'),
    ('ingresos diferidos', 'account'),
    ('gastos diferidos', 'account'),
    ('conciliar', 'account'),
    ('sepa', 'account'),
    ('pago sepa', 'account'),
    ('saldo bancario', 'account'),
    ('ajuste de divisas', 'account'),
    ('calendario fiscal', 'account'),
    ('cierre fiscal', 'account'),
    ('recalcular impuestos', 'account'),
    ('patrimonio neto', 'account'),
    ('cuenta bancaria', 'account'),
    ('cuenta: sincronizar diario en línea', 'account'),
    ('transferencias automáticas de cuenta', 'account'),
    ('facturación automática', 'account'),
    ('inventario', 'stock'),
    ('stock', 'stock'),
    ('reposición', 'stock'),
    ('replenishment', 'stock'),
    ('abastecimiento', 'stock'),
    ('fecha de caducidad', 'stock'),
    ('bloquear operaciones', 'stock'),
    ('punto de pedido', 'stock'),
    ('disponibilidad de producto', 'stock'),
    ('purchase', 'purchase'),
    ('pedido de compra', 'purchase'),
    ('solicitud de presupuesto', 'purchase'),
    ('suscripción', 'sale_subscription'),
    ('subscription', 'sale_subscription'),
    ('renovación', 'sale_subscription'),
    ('contrato vencido', 'sale_subscription'),
    ('pago recurrente', 'sale_subscription'),
    ('proyecto: crear tareas recurrentes', 'project'),
    ('proyecto: enviar calificación', 'project_rating'),
    ('project', 'project'),
    ('tarea vencida', 'project'),
    ('presupuesto de proyecto', 'project'),
    ('rentabilidad', 'project'),
    ('ausencias', 'hr_holidays'),
    ('holiday', 'hr_holidays'),
    ('asignación', 'hr_holidays'),
    ('nómina', 'hr_payroll'),
    ('payroll', 'hr_payroll'),
    ('evaluación', 'hr_appraisal'),
    ('appraisal', 'hr_appraisal'),
    ('hoja de asistencia', 'hr_attendance'),
    ('contrato de rr. hh.', 'hr_contract'),
    ('empleado de rr. hh.', 'hr'),
    ('permiso de trabajo', 'hr'),
    ('generar entradas de trabajo faltantes', 'hr_work_entry'),
    ('entradas de trabajo', 'hr_work_entry'),
    ('contrato del empleado', 'hr_contract'),
    ('ocr de gastos', 'hr_expense'),
    ('google calendar', 'google_calendar'),
    ('outlook: sincronización', 'microsoft_calendar'),
    ('calendario: recordatorio', 'calendar'),
    ('recordatorio de evento', 'calendar'),
    ('usuarios: notificar', 'base'),
    ('usuarios no registrados', 'base'),
    ('social: hacer las publicaciones', 'social'),
    ('publicaciones programadas', 'social'),
    ('crm: enriquecer leads', 'crm'),
    ('puntuación predictiva de leads', 'crm_lead_scoring'),
    ('probabilidades automatizadas', 'crm_lead_scoring'),
    ('lead/oportunidad', 'crm'),
    ('pago: transacciones posprocesadas', 'payment'),
    ('transacción de pago', 'payment'),
    ('autocompletar contacto', 'partner_autocomplete'),
    ('sincronización para autocompletar', 'partner_autocomplete'),
    ('ppc', 'website_sale'),
    ('affiliate', 'affiliate'),
    ('e-learning', 'website_slides'),
    ('elearning', 'website_slides'),
    ('canal de e-learning', 'website_slides'),
    ('slides', 'website_slides'),
    ('courses content', 'website_slides'),
    ('curso abierto', 'website_slides'),
    ('foro', 'website_forum'),
    ('forum', 'website_forum'),
    ('punto de venta', 'point_of_sale'),
    ('pos session', 'point_of_sale'),
    ('loan', 'account_loan'),
    ('préstamo', 'account_loan'),
    ('spread', 'account_spread_cost_revenue'),
    ('asset', 'account_asset'),
    ('amortización', 'account_asset'),
    ('activo fijo', 'account_asset'),
    ('intercompany', 'account_intercompany'),
    ('sale order', 'sale'),
    ('payment token', 'payment'),
    ('token de pago', 'payment'),
    ('importe a pagar', 'account'),
    ('tíquet de asistencia', 'helpdesk'),
    ('ticket', 'helpdesk'),
    ('evento: planificador de correo', 'event'),
    ('envío automático de correos de eventos', 'event'),
    ('evento:', 'event'),
]

# Crons OCA conocidos (palabras clave -> módulo OCA)
OCA_KEYWORDS_MAP = [
    ('cola de trabajos', 'queue_job'),
    ('vaciado automático de la cola', 'queue_job'),
    ('recolector de basura de trabajos', 'queue_job'),
    ('informe mis', 'mis_builder'),
    ('vacío de informes temporales', 'mis_builder'),
    ('instancia de informe mis', 'mis_builder'),
    ('fees invoice cron', 'openeducat_admission_enterprise'),
    ('factura de tarifas cron', 'openeducat_admission_enterprise'),
    ('detalles de las tarifas de los estudiantes', 'openeducat_admission_enterprise'),
    ('recordatorio materiales', 'openeducat_core'),
]

def classify_native(nombre, modelo):
    """Intenta clasificar como nativo Odoo por heurística de palabras clave."""
    text = (nombre + ' ' + modelo).lower()
    for keyword, module in NATIVE_KEYWORDS:
        if keyword in text:
            return module
    return None

def classify_oca(nombre, modelo):
    """Intenta clasificar como OCA/terceros por heurística de palabras clave."""
    text = (nombre + ' ' + modelo).lower()
    for keyword, module in OCA_KEYWORDS_MAP:
        if keyword in text:
            return module
    return None

# ============================================================
# PASO 1: Cargar xlsx
# ============================================================
print("Cargando xlsx...", flush=True)
wb = openpyxl.load_workbook(XLSX_PATH)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

actions = []
for row in ws.iter_rows(min_row=2, values_only=True):
    data = dict(zip(headers, row))
    actions.append({
        'activo': bool(data.get('Activo')),
        'modelo': str(data.get('Modelo') or ''),
        'nombre': str(data.get('Nombre de la acción') or ''),
        'intervalo': f"{data.get('Número de intervalos')} {data.get('Unidad de intervalo')}",
        'siguiente_ejecucion': str(data.get('Siguiente fecha de ejecución', '')),
        'prioridad': data.get('Prioridad'),
        'num_ejecuciones': data.get('Número de ejecuciones'),
    })

print(f"  {len(actions)} acciones cargadas", flush=True)

# ============================================================
# PASO 2: Construir índice global de XML (una sola llamada grep)
# ============================================================
print("Indexando XML de ir.cron en addons-extra/...", flush=True)

# Obtener lista de todos los ficheros XML con ir.cron
result = subprocess.run(
    ['grep', '-r', '--include=*.xml', '-l', 'ir.cron', 'addons-extra'],
    capture_output=True, text=True, cwd=BASE_DIR
)
cron_xml_files = [f.strip() for f in result.stdout.splitlines() if f.strip()]
print(f"  Ficheros XML con ir.cron: {len(cron_xml_files)}", flush=True)

# Para cada fichero, extraer los nombres de ir.cron definidos
# Patrón: <field name="name">TEXTO</field> dentro de <record model="ir.cron"
xml_cron_index = {}  # nombre_accion -> filepath (relativo)

for fpath in cron_xml_files:
    full_path = os.path.join(BASE_DIR, fpath)
    try:
        with open(full_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        # Buscar bloques ir.cron
        # Extraer todos los <field name="name">...</field>
        names_in_file = re.findall(r'<field\s+name=["\']name["\']>([^<]+)</field>', content)
        for name in names_in_file:
            name_clean = name.strip()
            if name_clean:
                xml_cron_index[name_clean] = fpath
    except Exception as e:
        pass

print(f"  Crons indexados en XML: {len(xml_cron_index)}", flush=True)

# ============================================================
# PASO 3: Construir índice global de _name Python en custom/OCA
# ============================================================
print("Indexando _name en Python files de addons-extra/...", flush=True)

result_py = subprocess.run(
    ['grep', '-r', '--include=*.py', '-n', '_name\s*=', 'addons-extra'],
    capture_output=True, text=True, cwd=BASE_DIR
)
py_name_index = {}  # model_name_technical -> filepath

for line in result_py.stdout.splitlines():
    # Formato: path/file.py:123:    _name = 'some.model'
    m = re.match(r'^(addons-extra/[^:]+):(\d+):\s*_name\s*=\s*[\'"]([^\'\"]+)[\'"]', line)
    if m:
        filepath, lineno, model_name = m.group(1), m.group(2), m.group(3)
        py_name_index[model_name.strip()] = filepath

print(f"  Modelos Python indexados: {len(py_name_index)}", flush=True)

# ============================================================
# FUNCIONES DE APOYO
# ============================================================
def get_module_info(filepath):
    """Sube por el árbol hasta encontrar __manifest__.py. filepath es relativo a BASE_DIR."""
    parts = filepath.split('/')
    for i in range(len(parts) - 1, 0, -1):
        candidate = os.path.join(BASE_DIR, *parts[:i])
        if os.path.exists(os.path.join(candidate, '__manifest__.py')):
            modulo = parts[i - 1]
            categoria = parts[i - 2] if i >= 2 else 'desconocida'
            return {
                'categoria': categoria,
                'modulo': modulo,
                'fichero': filepath,
            }
    return None

def classify_origen_by_category(categoria):
    if categoria in CUSTOM_CATEGORIES:
        return 'custom'
    elif categoria in OCA_CATEGORIES:
        return 'oca'
    return 'no_identificado'

# ============================================================
# PASO 4: Vincular cada acción
# ============================================================
print("Vinculando acciones...", flush=True)

findings = []

for action in actions:
    nombre = action['nombre']
    modelo = action['modelo']
    
    entry = {
        'nombre': nombre,
        'activo': action['activo'],
        'modelo': modelo,
        'intervalo': action['intervalo'],
        'siguiente_ejecucion': action['siguiente_ejecucion'],
        'origen': 'no_identificado',
        'categoria': '',
        'modulo': '',
        'fichero_xml': '',
        'confianza': 'baja',
        'notas': '',
    }
    
    # ---- ESTRATEGIA A: nombre exacto en índice XML ----
    if nombre in xml_cron_index:
        fpath = xml_cron_index[nombre]
        mod_info = get_module_info(fpath)
        if mod_info:
            cat = mod_info['categoria']
            entry['origen'] = classify_origen_by_category(cat)
            if entry['origen'] == 'no_identificado':
                # Si no está en categorías conocidas, pero está en addons-extra, es OCA o custom no identificado
                if fpath.startswith('addons-extra/'):
                    entry['origen'] = 'oca'  # fallback conservador
            entry['categoria'] = cat
            entry['modulo'] = mod_info['modulo']
            entry['fichero_xml'] = fpath
            entry['confianza'] = 'alta'
            findings.append(entry)
            continue
    
    # ---- ESTRATEGIA A2: búsqueda parcial de nombre en índice XML (para renombrados) ----
    nombre_lower = nombre.lower()
    partial_match = None
    for xml_name, fpath in xml_cron_index.items():
        # Match si el nombre del xlsx contiene el nombre XML o viceversa (mínimo 10 chars)
        xml_lower = xml_name.lower()
        if len(xml_lower) >= 10 and (xml_lower in nombre_lower or nombre_lower in xml_lower):
            partial_match = (xml_name, fpath)
            break
    
    if partial_match:
        xml_name, fpath = partial_match
        mod_info = get_module_info(fpath)
        if mod_info:
            cat = mod_info['categoria']
            entry['origen'] = classify_origen_by_category(cat)
            if entry['origen'] == 'no_identificado' and fpath.startswith('addons-extra/'):
                entry['origen'] = 'oca'
            entry['categoria'] = cat
            entry['modulo'] = mod_info['modulo']
            entry['fichero_xml'] = fpath
            entry['confianza'] = 'media'
            entry['notas'] = f'Match parcial con: "{xml_name}"'
            findings.append(entry)
            continue
    
    # ---- ESTRATEGIA B: modelo en índice Python ----
    # Intentar match directo del modelo con _name técnico
    # Primero intentar match por palabras clave del modelo en los _name conocidos
    modelo_words = re.findall(r'[a-z]{4,}', modelo.lower())
    py_match = None
    for tech_name, fpath in py_name_index.items():
        tech_lower = tech_name.lower().replace('.', ' ').replace('_', ' ')
        if any(w in tech_lower for w in modelo_words if len(w) >= 5):
            py_match = (tech_name, fpath)
            break
    
    if py_match:
        tech_name, fpath = py_match
        mod_info = get_module_info(fpath)
        if mod_info:
            cat = mod_info['categoria']
            origen = classify_origen_by_category(cat)
            if origen in ('custom', 'oca'):
                entry['origen'] = origen
                entry['categoria'] = cat
                entry['modulo'] = mod_info['modulo']
                entry['confianza'] = 'media'
                entry['notas'] = f'Match por modelo "{tech_name}" en {fpath}'
                findings.append(entry)
                continue
    
    # ---- ESTRATEGIA C: heurística nativa Odoo ----
    native_mod = classify_native(nombre, modelo)
    if native_mod:
        entry['origen'] = 'nativo_odoo'
        entry['modulo'] = native_mod
        entry['confianza'] = 'media'
        findings.append(entry)
        continue
    
    # ---- ESTRATEGIA C2: heurística OCA/Terceros ----
    oca_mod = classify_oca(nombre, modelo)
    if oca_mod:
        entry['origen'] = 'oca'
        entry['modulo'] = oca_mod
        entry['confianza'] = 'media'
        entry['categoria'] = 'terceros'
        findings.append(entry)
        continue
    
    # Sin clasificar
    entry['notas'] = f'Sin match. Modelo="{modelo}"'
    findings.append(entry)

# ============================================================
# PASO 5: Estadísticas
# ============================================================
total = len(findings)
activas = sum(1 for f in findings if f['activo'])
inactivas = total - activas
custom_count = sum(1 for f in findings if f['origen'] == 'custom')
oca_count = sum(1 for f in findings if f['origen'] == 'oca')
nativo_count = sum(1 for f in findings if f['origen'] == 'nativo_odoo')
no_id_count = sum(1 for f in findings if f['origen'] == 'no_identificado')

print(f"\n=== ESTADÍSTICAS ===")
print(f"Total: {total}")
print(f"Activas: {activas} | Inactivas: {inactivas}")
print(f"Custom IRG/ISEP: {custom_count}")
print(f"OCA / Terceros: {oca_count}")
print(f"Nativo Odoo: {nativo_count}")
print(f"No identificadas: {no_id_count}")

# ============================================================
# PASO 6: Guardar JSON
# ============================================================
with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(findings, f, ensure_ascii=False, indent=2, default=str)
print(f"\nJSON guardado en: {OUTPUT_JSON}")

# ============================================================
# PASO 7: Generar _findings_crons.md
# ============================================================
today = datetime.today().strftime('%Y-%m-%d')

lines = []
lines.append(f"# Findings: Acciones Planificadas (ir.cron) — {today}\n")
lines.append(f"Análisis del servidor de producción Odoo 16 IRG/ISEP.\n")
lines.append(f"Fuente: `doc/csv-analizer/Acciones planificadas (ir.cron).xlsx`\n")
lines.append("")
lines.append("## Resumen\n")
lines.append(f"| Métrica | Valor |")
lines.append(f"|---|---|")
lines.append(f"| Total acciones analizadas | **{total}** |")
lines.append(f"| Activas | **{activas}** |")
lines.append(f"| Inactivas | **{inactivas}** |")
lines.append(f"| Custom IRG/ISEP (alta confianza) | **{sum(1 for f in findings if f['origen']=='custom' and f['confianza']=='alta')}** |")
lines.append(f"| Custom IRG/ISEP (media confianza) | **{sum(1 for f in findings if f['origen']=='custom' and f['confianza']=='media')}** |")
lines.append(f"| OCA / Terceros | **{oca_count}** |")
lines.append(f"| Nativo Odoo | **{nativo_count}** |")
lines.append(f"| No identificadas | **{no_id_count}** |")
lines.append("")
lines.append("## Mapa completo\n")
lines.append("| # | Nombre de la acción | Activo | Modelo | Intervalo | Origen | Módulo | Categoría | Fichero XML | Confianza |")
lines.append("|---|---|---|---|---|---|---|---|---|---|")

for i, f in enumerate(findings, 1):
    activo_str = "✅" if f['activo'] else "❌"
    origen_icon = {
        'custom': '🔧',
        'oca': '📦',
        'nativo_odoo': '🏠',
        'no_identificado': '❓',
    }.get(f['origen'], '❓')
    fichero = f['fichero_xml'] if f['fichero_xml'] else '—'
    # Acortar filepath para legibilidad
    if fichero and len(fichero) > 60:
        fichero = '...' + fichero[-57:]
    lines.append(
        f"| {i} | {f['nombre']} | {activo_str} | {f['modelo']} | {f['intervalo']} "
        f"| {origen_icon} {f['origen']} | {f['modulo']} | {f['categoria']} | {fichero} | {f['confianza']} |"
    )

lines.append("")
lines.append("## Acciones no identificadas\n")
no_id_list = [f for f in findings if f['origen'] == 'no_identificado']
if no_id_list:
    lines.append("Acciones para las que no se encontró módulo fuente:\n")
    for f in no_id_list:
        lines.append(f"- **{f['nombre']}** (Modelo: `{f['modelo']}`) — {f['notas']}")
else:
    lines.append("_Ninguna acción sin identificar._")

lines.append("")
lines.append("## Custom IRG/ISEP — detalle\n")
custom_list = [f for f in findings if f['origen'] == 'custom']
if custom_list:
    lines.append("| Nombre de la acción | Módulo | Categoría | Fichero XML | Confianza |")
    lines.append("|---|---|---|---|---|")
    for f in custom_list:
        fichero = f['fichero_xml'] if f['fichero_xml'] else '—'
        if fichero and len(fichero) > 70:
            fichero = '...' + fichero[-67:]
        lines.append(f"| {f['nombre']} | {f['modulo']} | {f['categoria']} | {fichero} | {f['confianza']} |")
else:
    lines.append("_Ninguna acción custom identificada._")

lines.append("")
lines.append("## OCA / Terceros — detalle\n")
oca_list = [f for f in findings if f['origen'] == 'oca']
if oca_list:
    lines.append("| Nombre de la acción | Módulo | Categoría | Fichero XML | Confianza |")
    lines.append("|---|---|---|---|---|")
    for f in oca_list:
        fichero = f['fichero_xml'] if f['fichero_xml'] else '—'
        if fichero and len(fichero) > 70:
            fichero = '...' + fichero[-67:]
        lines.append(f"| {f['nombre']} | {f['modulo']} | {f['categoria']} | {fichero} | {f['confianza']} |")
else:
    lines.append("_Ninguna acción OCA identificada._")

lines.append("")
lines.append("## Nativo Odoo — detalle\n")
nativo_list = [f for f in findings if f['origen'] == 'nativo_odoo']
if nativo_list:
    lines.append("| Nombre de la acción | Módulo Odoo | Confianza |")
    lines.append("|---|---|---|")
    for f in nativo_list:
        lines.append(f"| {f['nombre']} | {f['modulo']} | {f['confianza']} |")
else:
    lines.append("_Ninguna acción nativa identificada._")

# Escribir el fichero
with open(OUTPUT_MD, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines) + '\n')

print(f"Markdown guardado en: {OUTPUT_MD}")
print("\n¡Análisis completado!")
